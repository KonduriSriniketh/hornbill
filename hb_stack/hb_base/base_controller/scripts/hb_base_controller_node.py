#!/usr/bin/env python3

import math
import time
import openpyxl
import rospy
import tf
import orienbus as orien
from nav_msgs.msg import Odometry
from hb_msgs.msg import Speed
from geometry_msgs.msg import Quaternion
from geometry_msgs.msg import Twist, PoseWithCovarianceStamped
from sensor_msgs.msg import Imu

class BaseController(object):

    def __init__(self):

        self._wheel_base     = rospy.get_param('wheel_base', 0.14)
        self._wheel_radius   = rospy.get_param('wheel_radius', 0.03)
        self._port           = rospy.get_param('port', '/dev/ttyUSB0')
        self._port_arm       = rospy.get_param('port_arm', '/dev/ttyUSB1')
        self._left_motor_id  = rospy.get_param('left_motor_id')
        self._right_motor_id = rospy.get_param('right_motor_id')
        self._arm_motor_id   = rospy.get_param('arm_motor_id')
        self._gear_ratio     = rospy.get_param('gear_ratio')
        self.max_RPM         = rospy.get_param('max_RPM', 4000)

        # Create orienbus object with port name
        orienbus_right = orien.OrienBus(self._port)
        orienbus_left = orien.OrienBus(self._port)

        orienbus_arm = orien.OrienBus(self._port_arm)

        # Initialze motors with slave addresses
        self._left_motor = orienbus_left.initialize(self._left_motor_id)
        self._right_motor = orienbus_right.initialize(self._right_motor_id)
        self._arm_motor = orienbus_arm.initialize(self._arm_motor_id)

        self.wheels_speed_pub = rospy.Publisher('/wheels_speed', Speed, queue_size=1)
        rospy.Subscriber("/cmd_vel", Twist, self.cmdVelCallback)

        
        self.linear_x=0
        self.angular_z=0
        self.linear_y=0

        self.left_wheel_speed = 0
        self.right_wheel_speed = 0

        self.arm_speed=0

        self.current_time = rospy.Time.now()
        self.prev_time = rospy.Time.now()
        self.cb_timer = rospy.Time.now()
        self.prev_yaw = 0

        # Temporary Variables
        self.right_wheel_speed_rpm = 0
        self.left_wheel_speed_rpm = 0
        self.l = 0
        self.r = 0

	
            
    def cmdVelCallback(self, msg):

        self.linear_x = msg.linear.x
        self.linear_y = msg.linear.y
        self.angular_z = msg.angular.z

        self.left_wheel_speed  = (self.linear_x / self._wheel_radius) - (self._wheel_base * self.angular_z/(2*self._wheel_radius) )
        self.right_wheel_speed = (self.linear_x / self._wheel_radius) + (self._wheel_base * self.angular_z/(2*self._wheel_radius) )
        self.arm_speed = self.linear_y
        

        self.cb_timer = rospy.Time.now()

    def PublishWheelSpeed(self):

        speed = Speed()
        l_speed_rpm = self._left_motor.readSpeed() 
        r_speed_rpm = - self._right_motor.readSpeed() 
        if (l_speed_rpm > 30000 or l_speed_rpm < -30000 or r_speed_rpm >30000 or r_speed_rpm < -30000):
            self.l = self.l
            self.r = self.r
            print("nosie value found" )
            print("left noise =", l_speed_rpm)
            print("right noise =", r_speed_rpm)

        else:
            self.l = l_speed_rpm
            self.r = r_speed_rpm
        
        print("read l_speed_rpm internal = ", self.l)
        print("read r_speed_rpm internal = ", self.r)

        self.l_speed_rpm = self.l / self._gear_ratio # rpm
        self.r_speed_rpm = self.r / self._gear_ratio

        speed.wheel_left = self.l_speed_rpm * 2 * math.pi / 60 # converting to rad/sec
        speed.wheel_right = self.r_speed_rpm * 2 * math.pi / 60

        print("read loop left  wheel angular speed rad/sec = ", speed.wheel_left)
        print("read loop right wheel angluar speed rad/sec = ", speed.wheel_right)
        print("---")
        self.wheels_speed_pub.publish(speed)

    def WriteWheelSpeed(self,right_wheel_speed, left_wheel_speed):
        
        print("write left  wheel angular speed = ", left_wheel_speed)
        print("write right wheel angular speed = ", right_wheel_speed)
        right_wheel_speed_rpm = right_wheel_speed * 60/ (2*math.pi) * self._gear_ratio #rpm internal
        left_wheel_speed_rpm = left_wheel_speed * 60/ (2*math.pi) * self._gear_ratio
        
        self.left_wheel_speed_rpm = int(left_wheel_speed_rpm)
        self.right_wheel_speed_rpm = int(right_wheel_speed_rpm)
        print("write left  wheel speed rpm = ", int(left_wheel_speed_rpm))
        print("write right wheel speed rpm = ", int(right_wheel_speed_rpm))
        # print("---")

        self._left_motor.writeSpeed(int(left_wheel_speed_rpm))
        self._right_motor.writeSpeed(int(-right_wheel_speed_rpm))


    def WriteArmSpeed(self,arm_speed):

        if arm_speed > 0:
            self.arm_speed = 200
        elif (arm_speed < 0):
             self.arm_speed = -200
        else:
            self.arm_speed = 0

        self._arm_motor.writeSpeed(self.arm_speed)


def main():
    rospy.init_node("base_controller_node", anonymous=True)

    file_path = '/home/sutd/hb_ws/src/hb_stack/hb_base/base_controller/excel/speeds.xlsx'

    _object = BaseController()

    rate = rospy.Rate(10)
    ctrl_c = False

    def shutdownhook():
        ctrl_c = True

    rospy.on_shutdown(shutdownhook)

    # try:
    #     workbook = openpyxl.load_workbook(file_path)
    # except FileNotFoundError:
    #     # If the file doesn't exist, create a new workbook
    #     workbook = openpyxl.Workbook()

    # # Select the active sheet or create a new one
    # sheet = workbook.active
    # # Start writing from the first available empty row
    # start_row = sheet.max_row + 1  # Next empty row

    # # Set headers for columns: time, x, y, value, and time
    # if sheet.max_row == 1:  # Check if the sheet is empty (only one row means it's empty)
    #     sheet['A1'] = 'rostime'
    #     sheet['B1'] = 'left_write_raw_speed_data'
    #     sheet['C1'] = 'right_write_raw_speed_data'  
    #     sheet['D1'] = 'left_read_raw_speed_data_internal'
    #     sheet['E1'] = 'right_read_raw_speed_data'

    # # Start writing from the first available empty row after headers
    # start_row = sheet.max_row + 1  # Next empty row

    while not rospy.is_shutdown():

        _object.current_time = rospy.Time.now()
        _object.PublishWheelSpeed()
        _object.WriteWheelSpeed(_object.right_wheel_speed, _object.left_wheel_speed)
        _object.WriteArmSpeed(_object.arm_speed)
        # Write the data
        # sheet[f'A{start_row}'] = _object.current_time.to_sec() 
        # sheet[f'B{start_row}'] = _object.left_wheel_speed_rpm  
        # sheet[f'C{start_row}'] = _object.right_wheel_speed_rpm  
        # sheet[f'D{start_row}'] = _object.l  
        # sheet[f'E{start_row}'] = _object.r

        # # Increment the row for the next input
        # start_row += 1
        # workbook.save(file_path)

        rate.sleep()

if __name__ == '__main__':
    main()
